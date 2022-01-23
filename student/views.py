
from django.http import HttpResponse
from .models import grades
from django.contrib.auth.models import User
from .serializer import gradeserializer
from rest_framework import generics
from rest_framework.views import APIView
from openpyxl import Workbook, load_workbook 
from openpyxl.utils import get_column_letter
import xlsxwriter
from rest_framework.response import Response 
from rest_framework.authtoken.models import Token
import xlrd
from rest_framework.authentication import TokenAuthentication
from rest_framework.permissions import IsAuthenticated
from .serializer import userSerializer
# Create your views here.

class UserRegister(APIView):
    def post(self, request):
        ser =userSerializer(data=request.data)
        ser.is_valid(raise_exception=True)
        if not ser.is_valid():
            return Response({'status': 403 ,'error':ser.errors,'data':"something is wrong"})
        ser.save()
        user=User.objects.get(username=ser.data['username'])
        token, created = Token.objects.get_or_create(user=user)
        return Response({
            'token': token.key,
            'user_id': user.pk,
            'email': user.email
        })

class getToken(APIView):
    authentication_classes =[TokenAuthentication]
    permission_classes=[IsAuthenticated]
    def post(self,request,id):
        
        user=User.objects.get(id=id)
        token = Token.objects.get(user=user)
        return Response({
            'token': token.key,
            'user_id': user.pk,
            'email': user.email
        })
        

class GradeAPI(generics.ListCreateAPIView):
    queryset=grades.objects.all()
    serializer_class=gradeserializer
#88f75bff2e9bf8b4f66eea7d278061023c6e720e

class ExcelToDatabase(generics.CreateAPIView):
    authentication_classes =[TokenAuthentication]
    permission_classes=[IsAuthenticated]
    queryset=grades.objects.all()
    serializer_class=gradeserializer
    def get(self, request, *args, **kwargs):
        data=grades.objects.all()
        #print(data[0].name)
        #print()
        wb = load_workbook("grade.xlsx")
        ws = wb.active
        '''for row in range(2,data.count()): 
            for col in range(1,5):
                print(ws[str(get_column_letter(col))+str(row)].value)'''
        for row in range (2,ws.max_row+1):
                
                data=grades.objects.create(
                     name=ws['A'+str(row)].value,
                    maths=ws['B'+str(row)].value,
                    science=ws['C'+str(row)].value,
                    english=ws['D'+str(row)].value
                    )
        return HttpResponse('done',row)

class DatabaseToExcel(generics.ListAPIView):
    authentication_classes =[TokenAuthentication]
    permission_classes=[IsAuthenticated]
    queryset=grades.objects.all()
    serializer_class=gradeserializer
    def get(self, request, *args, **kwargs):
        data=grades.objects.all()
        count=grades.objects.count()
        #print(count)
        wb=xlsxwriter.Workbook('grade.xlsx')
        ws=wb.add_worksheet()
        ws.write('A1','NAME')
        ws.write('B1','MATHS')
        ws.write('C1','SCIENCE')
        ws.write('D1','ENGLISH')
        for i in range(2,count+2):
            ws.write('A'+str(i),data[i-2].name)
            ws.write('B'+str(i),data[i-2].maths)
            ws.write('C'+str(i),data[i-2].science)
            ws.write('D'+str(i),data[i-2].english)
        wb.close()
        return HttpResponse('done')